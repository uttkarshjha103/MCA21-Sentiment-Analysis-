"""
Report generation service for Excel and CSV exports.
"""
import csv
import io
import logging
import os
import uuid
from datetime import datetime
from typing import Any, Dict, List, Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from motor.motor_asyncio import AsyncIOMotorDatabase

from ..models.report import ReportCreate, ReportFilters, ReportResponse

logger = logging.getLogger(__name__)

REPORTS_DIR = os.path.join("uploads", "reports")


def _ensure_reports_dir():
    os.makedirs(REPORTS_DIR, exist_ok=True)


class ReportService:
    def __init__(self, db: AsyncIOMotorDatabase):
        self.db = db
        self.collection = db.reports

    def _build_comment_filter(self, filters: Optional[ReportFilters]) -> Dict[str, Any]:
        query: Dict[str, Any] = {}
        if not filters:
            return query
        if filters.date_from or filters.date_to:
            query["date_submitted"] = {}
            if filters.date_from:
                query["date_submitted"]["$gte"] = filters.date_from
            if filters.date_to:
                query["date_submitted"]["$lte"] = filters.date_to
        if filters.sentiment:
            query["sentiment.label"] = filters.sentiment
        if filters.language:
            query["original_language"] = filters.language
        if filters.source:
            query["source"] = filters.source
        if filters.keywords:
            query["keywords.text"] = {"$in": filters.keywords}
        return query

    async def _fetch_comments(self, filters: Optional[ReportFilters]) -> List[Dict[str, Any]]:
        query = self._build_comment_filter(filters)
        cursor = self.db.comments.find(query).sort("date_submitted", -1)
        comments = []
        async for doc in cursor:
            doc["_id"] = str(doc["_id"])
            comments.append(doc)
        return comments

    async def _compute_metadata(self, comments: List[Dict[str, Any]], filters: Optional[ReportFilters]) -> Dict[str, Any]:
        total = len(comments)
        sentiment_dist: Dict[str, int] = {}
        language_dist: Dict[str, int] = {}
        for c in comments:
            label = (c.get("sentiment") or {}).get("label", "unknown")
            sentiment_dist[label] = sentiment_dist.get(label, 0) + 1
            lang = c.get("original_language") or "unknown"
            language_dist[lang] = language_dist.get(lang, 0) + 1
        return {
            "generated_at": datetime.utcnow().isoformat(),
            "system": "MCA21 Sentiment Analysis System",
            "total_comments": total,
            "sentiment_distribution": sentiment_dist,
            "language_distribution": language_dist,
            "filters_applied": filters.dict() if filters else {},
        }

    def _generate_excel_bytes(self, comments: List[Dict[str, Any]], metadata: Dict[str, Any], title: str) -> bytes:
        wb = openpyxl.Workbook()
        meta_ws = wb.active
        meta_ws.title = "Metadata"
        meta_ws.append(["MCA21 Sentiment Analysis Report"])
        meta_ws.append(["Title", title])
        for key, value in metadata.items():
            meta_ws.append([key, str(value)])
        meta_ws.column_dimensions["A"].width = 30
        meta_ws.column_dimensions["B"].width = 60

        data_ws = wb.create_sheet("Comments")
        white_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4472C4")
        columns = ["Comment ID", "Comment Text", "Source", "Date Submitted", "Language", "Sentiment", "Confidence", "Keywords"]
        data_ws.append(columns)
        for col_idx in range(1, len(columns) + 1):
            cell = data_ws.cell(row=1, column=col_idx)
            cell.font = white_font
            cell.fill = header_fill

        for comment in comments:
            sentiment = comment.get("sentiment") or {}
            keywords_list = comment.get("keywords") or []
            keyword_texts = ", ".join(k.get("text", "") for k in keywords_list[:10])
            date_submitted = comment.get("date_submitted")
            if isinstance(date_submitted, datetime):
                date_submitted = date_submitted.isoformat()
            data_ws.append([
                comment.get("_id", ""),
                comment.get("comment_text", ""),
                comment.get("source", ""),
                date_submitted or "",
                comment.get("original_language", ""),
                sentiment.get("label", ""),
                sentiment.get("confidence_score", ""),
                keyword_texts,
            ])

        for col_idx in range(1, len(columns) + 1):
            data_ws.column_dimensions[get_column_letter(col_idx)].width = 20
        data_ws.column_dimensions["B"].width = 60

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def _generate_csv_bytes(self, comments: List[Dict[str, Any]], metadata: Dict[str, Any], include_metadata: bool) -> bytes:
        buf = io.StringIO()
        writer = csv.writer(buf)
        if include_metadata:
            writer.writerow(["# MCA21 Sentiment Analysis Export"])
            for key, value in metadata.items():
                writer.writerow([f"# {key}", str(value)])
            writer.writerow([])
        writer.writerow(["comment_id", "comment_text", "source", "date_submitted", "language", "sentiment_label", "confidence_score", "keywords"])
        for comment in comments:
            sentiment = comment.get("sentiment") or {}
            keywords_list = comment.get("keywords") or []
            keyword_texts = "|".join(k.get("text", "") for k in keywords_list[:10])
            date_submitted = comment.get("date_submitted")
            if isinstance(date_submitted, datetime):
                date_submitted = date_submitted.isoformat()
            writer.writerow([
                comment.get("_id", ""),
                comment.get("comment_text", ""),
                comment.get("source", ""),
                date_submitted or "",
                comment.get("original_language", ""),
                sentiment.get("label", ""),
                sentiment.get("confidence_score", ""),
                keyword_texts,
            ])
        return buf.getvalue().encode("utf-8")

    async def _create_report_record(self, report_id: str, title: str, report_type: str, user_id: str, filters: Optional[ReportFilters]) -> None:
        doc = {
            "report_id": report_id,
            "title": title,
            "report_type": report_type,
            "status": "pending",
            "generated_by": user_id,
            "generated_at": datetime.utcnow(),
            "completed_at": None,
            "file_path": None,
            "filters_applied": filters.dict() if filters else {},
            "metadata": None,
            "error_message": None,
        }
        await self.collection.insert_one(doc)

    async def _update_report_record(self, report_id: str, update: Dict[str, Any]) -> None:
        await self.collection.update_one({"report_id": report_id}, {"$set": update})

    async def get_report_status(self, report_id: str) -> Optional[ReportResponse]:
        doc = await self.collection.find_one({"report_id": report_id})
        if not doc:
            return None
        doc.pop("_id", None)
        return ReportResponse(**doc)

    async def generate_excel_report(self, request: ReportCreate, user_id: str) -> ReportResponse:
        _ensure_reports_dir()
        report_id = str(uuid.uuid4())
        await self._create_report_record(report_id, request.title, "excel", user_id, request.filters)
        try:
            await self._update_report_record(report_id, {"status": "processing"})
            comments = await self._fetch_comments(request.filters)
            metadata = await self._compute_metadata(comments, request.filters)
            excel_bytes = self._generate_excel_bytes(comments, metadata, request.title)
            filename = f"report_{report_id}.xlsx"
            file_path = os.path.join(REPORTS_DIR, filename)
            with open(file_path, "wb") as f:
                f.write(excel_bytes)
            await self._update_report_record(report_id, {
                "status": "completed",
                "completed_at": datetime.utcnow(),
                "file_path": file_path,
                "metadata": metadata,
            })
        except Exception as exc:
            logger.error(f"Excel report generation failed: {exc}")
            await self._update_report_record(report_id, {"status": "failed", "error_message": str(exc)})

        doc = await self.collection.find_one({"report_id": report_id})
        doc.pop("_id", None)
        response = ReportResponse(**doc)
        if response.status == "completed":
            response.download_url = f"/api/v1/reports/{report_id}/download"
        return response

    async def export_csv(self, filters: Optional[ReportFilters], include_metadata: bool = True) -> bytes:
        comments = await self._fetch_comments(filters)
        metadata = await self._compute_metadata(comments, filters)
        return self._generate_csv_bytes(comments, metadata, include_metadata)

    async def get_report_file_path(self, report_id: str) -> Optional[str]:
        doc = await self.collection.find_one({"report_id": report_id})
        if not doc or doc.get("status") != "completed":
            return None
        return doc.get("file_path")
