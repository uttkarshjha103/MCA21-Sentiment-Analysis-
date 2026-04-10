"""
Unit tests for file upload functionality.
"""
import pytest
from io import BytesIO
from fastapi import UploadFile
from app.services.upload import UploadService
from app.core.exceptions import FileProcessingError


class TestUploadService:
    """Test cases for UploadService."""
    
    def test_validate_csv_extension_valid(self):
        """Test CSV extension validation with valid files."""
        assert UploadService.validate_file_extension("comments.csv", "csv")
        assert UploadService.validate_file_extension("data.txt", "csv")
    
    def test_validate_csv_extension_invalid(self):
        """Test CSV extension validation with invalid files."""
        with pytest.raises(FileProcessingError):
            UploadService.validate_file_extension("comments.xlsx", "csv")
        
        with pytest.raises(FileProcessingError):
            UploadService.validate_file_extension("data.pdf", "csv")
    
    def test_validate_excel_extension_valid(self):
        """Test Excel extension validation with valid files."""
        assert UploadService.validate_file_extension("comments.xlsx", "excel")
        assert UploadService.validate_file_extension("data.xls", "excel")
    
    def test_validate_excel_extension_invalid(self):
        """Test Excel extension validation with invalid files."""
        with pytest.raises(FileProcessingError):
            UploadService.validate_file_extension("comments.csv", "excel")
        
        with pytest.raises(FileProcessingError):
            UploadService.validate_file_extension("data.pdf", "excel")
    
    @pytest.mark.asyncio
    async def test_process_csv_upload_valid(self):
        """Test CSV upload with valid data."""
        csv_content = b"comment_text,source,date_submitted\nThis is a great policy,test,2024-01-15\nI disagree with this,test,2024-01-16"
        
        file = UploadFile(
            filename="test.csv",
            file=BytesIO(csv_content)
        )
        
        comments, errors = await UploadService.process_csv_upload(file)
        
        assert len(comments) == 2
        assert comments[0].comment_text == "This is a great policy"
        assert comments[1].comment_text == "I disagree with this"
        assert len(errors) == 0
    
    @pytest.mark.asyncio
    async def test_process_csv_upload_missing_column(self):
        """Test CSV upload with missing required column."""
        csv_content = b"text,source\nThis is a comment,test"
        
        file = UploadFile(
            filename="test.csv",
            file=BytesIO(csv_content)
        )
        
        with pytest.raises(FileProcessingError) as exc_info:
            await UploadService.process_csv_upload(file)
        
        assert "comment_text" in str(exc_info.value)
    
    @pytest.mark.asyncio
    async def test_process_csv_upload_empty_comments(self):
        """Test CSV upload with empty comment text."""
        csv_content = b"comment_text,source\n,test\n  ,test"
        
        file = UploadFile(
            filename="test.csv",
            file=BytesIO(csv_content)
        )
        
        with pytest.raises(FileProcessingError) as exc_info:
            await UploadService.process_csv_upload(file)
        
        assert "No valid comments" in str(exc_info.value)
    
    @pytest.mark.asyncio
    async def test_process_csv_upload_with_errors(self):
        """Test CSV upload with some invalid rows."""
        csv_content = b"comment_text,source,date_submitted\nValid comment,test,2024-01-15\n,test,2024-01-16\nAnother valid comment,test,invalid-date"
        
        file = UploadFile(
            filename="test.csv",
            file=BytesIO(csv_content)
        )
        
        comments, errors = await UploadService.process_csv_upload(file)
        
        assert len(comments) == 2
        assert len(errors) == 2  # One for empty comment, one for invalid date
    
    @pytest.mark.asyncio
    async def test_process_csv_upload_long_comment(self):
        """Test CSV upload with comment exceeding max length."""
        long_text = "a" * 10001
        csv_content = f"comment_text,source\n{long_text},test".encode()
        
        file = UploadFile(
            filename="test.csv",
            file=BytesIO(csv_content)
        )
        
        with pytest.raises(FileProcessingError) as exc_info:
            await UploadService.process_csv_upload(file)
        
        assert "No valid comments" in str(exc_info.value)
    
    @pytest.mark.asyncio
    async def test_process_csv_upload_invalid_format(self):
        """Test CSV upload with invalid CSV format."""
        csv_content = b"This is not a valid CSV format"
        
        file = UploadFile(
            filename="test.csv",
            file=BytesIO(csv_content)
        )
        
        with pytest.raises(FileProcessingError) as exc_info:
            await UploadService.process_csv_upload(file)
        
        assert "comment_text" in str(exc_info.value)
    
    @pytest.mark.asyncio
    async def test_process_excel_upload_valid(self):
        """Test Excel upload with valid data."""
        # Create a simple Excel file using openpyxl
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        ws.append(["comment_text", "source", "date_submitted"])
        ws.append(["This is a great policy", "test", "2024-01-15"])
        ws.append(["I disagree with this", "test", "2024-01-16"])
        
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        file = UploadFile(
            filename="test.xlsx",
            file=excel_buffer
        )
        
        comments, errors = await UploadService.process_excel_upload(file)
        
        assert len(comments) == 2
        assert comments[0].comment_text == "This is a great policy"
        assert comments[1].comment_text == "I disagree with this"
        assert len(errors) == 0
    
    @pytest.mark.asyncio
    async def test_process_excel_upload_missing_column(self):
        """Test Excel upload with missing required column."""
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        ws.append(["text", "source"])
        ws.append(["This is a comment", "test"])
        
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        file = UploadFile(
            filename="test.xlsx",
            file=excel_buffer
        )
        
        with pytest.raises(FileProcessingError) as exc_info:
            await UploadService.process_excel_upload(file)
        
        assert "comment_text" in str(exc_info.value)
    
    @pytest.mark.asyncio
    async def test_process_excel_upload_with_optional_fields(self):
        """Test Excel upload with optional fields."""
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        ws.append(["comment_text", "source", "original_language"])
        ws.append(["This is a comment", "test", "en"])
        ws.append(["Ceci est un commentaire", "test", "fr"])
        
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        file = UploadFile(
            filename="test.xlsx",
            file=excel_buffer
        )
        
        comments, errors = await UploadService.process_excel_upload(file)
        
        assert len(comments) == 2
        assert comments[0].original_language == "en"
        assert comments[1].original_language == "fr"
    
    @pytest.mark.asyncio
    async def test_process_excel_upload_invalid_format(self):
        """Test Excel upload with invalid file format."""
        invalid_content = b"This is not an Excel file"
        
        file = UploadFile(
            filename="test.xlsx",
            file=BytesIO(invalid_content)
        )
        
        with pytest.raises(FileProcessingError) as exc_info:
            await UploadService.process_excel_upload(file)
        
        assert "Invalid Excel format" in str(exc_info.value)
    
    @pytest.mark.asyncio
    async def test_process_csv_case_insensitive_columns(self):
        """Test CSV upload with case-insensitive column names."""
        csv_content = b"Comment_Text,SOURCE\nThis is a comment,test"
        
        file = UploadFile(
            filename="test.csv",
            file=BytesIO(csv_content)
        )
        
        comments, errors = await UploadService.process_csv_upload(file)
        
        assert len(comments) == 1
        assert comments[0].comment_text == "This is a comment"



class TestBulkInsertOperations:
    """Test cases for bulk insert operations with progress tracking."""
    
    @pytest.mark.asyncio
    async def test_bulk_insert_comments_success(self):
        """Test successful bulk insert of comments."""
        from mongomock_motor import AsyncMongoMockClient
        from app.models.comment import CommentCreate
        from bson import ObjectId
        
        # Create mock database
        client = AsyncMongoMockClient()
        db = client.test_db
        
        # Create test comments
        comments = [
            CommentCreate(
                comment_text=f"Test comment {i}",
                source="test_bulk"
            )
            for i in range(10)
        ]
        
        user_id = ObjectId()
        
        # Perform bulk insert
        result = await UploadService.bulk_insert_comments(
            comments=comments,
            user_id=user_id,
            db=db
        )
        
        assert result['success'] is True
        assert result['stored_count'] == 10
        assert result['failed_count'] == 0
        assert 'upload_id' in result
        
        # Verify comments were stored
        stored_comments = await db.comments.count_documents({})
        assert stored_comments == 10
        
        # Verify progress tracking was created
        progress = await db.upload_progress.find_one({"upload_id": result['upload_id']})
        assert progress is not None
        assert progress['status'] == 'completed'
        assert progress['stored_comments'] == 10
    
    @pytest.mark.asyncio
    async def test_bulk_insert_large_batch(self):
        """Test bulk insert with large number of comments."""
        from mongomock_motor import AsyncMongoMockClient
        from app.models.comment import CommentCreate
        from bson import ObjectId
        
        # Create mock database
        client = AsyncMongoMockClient()
        db = client.test_db
        
        # Create 250 test comments (more than batch size of 100)
        comments = [
            CommentCreate(
                comment_text=f"Test comment {i}",
                source="test_bulk"
            )
            for i in range(250)
        ]
        
        user_id = ObjectId()
        
        # Perform bulk insert
        result = await UploadService.bulk_insert_comments(
            comments=comments,
            user_id=user_id,
            db=db
        )
        
        assert result['success'] is True
        assert result['stored_count'] == 250
        assert result['total_comments'] == 250
        
        # Verify all comments were stored
        stored_comments = await db.comments.count_documents({})
        assert stored_comments == 250
    
    @pytest.mark.asyncio
    async def test_get_upload_progress(self):
        """Test retrieving upload progress."""
        from mongomock_motor import AsyncMongoMockClient
        from app.models.comment import UploadProgress
        from datetime import datetime
        
        # Create mock database
        client = AsyncMongoMockClient()
        db = client.test_db
        
        # Insert test progress record
        upload_id = "test_upload_123"
        progress_data = {
            "upload_id": upload_id,
            "user_id": "user_123",
            "total_comments": 100,
            "processed_comments": 50,
            "stored_comments": 48,
            "failed_comments": 2,
            "status": "processing",
            "errors": [],
            "created_at": datetime.utcnow(),
            "updated_at": datetime.utcnow()
        }
        
        await db.upload_progress.insert_one(progress_data)
        
        # Retrieve progress
        progress = await UploadService.get_upload_progress(upload_id, db)
        
        assert progress.upload_id == upload_id
        assert progress.total_comments == 100
        assert progress.processed_comments == 50
        assert progress.stored_comments == 48
        assert progress.status == "processing"
    
    @pytest.mark.asyncio
    async def test_get_upload_progress_not_found(self):
        """Test retrieving non-existent upload progress."""
        from mongomock_motor import AsyncMongoMockClient
        from app.core.exceptions import FileProcessingError
        
        # Create mock database
        client = AsyncMongoMockClient()
        db = client.test_db
        
        # Try to retrieve non-existent progress
        with pytest.raises(FileProcessingError) as exc_info:
            await UploadService.get_upload_progress("nonexistent_id", db)
        
        assert "not found" in str(exc_info.value)
    
    @pytest.mark.asyncio
    async def test_bulk_insert_with_custom_upload_id(self):
        """Test bulk insert with custom upload ID."""
        from mongomock_motor import AsyncMongoMockClient
        from app.models.comment import CommentCreate
        from bson import ObjectId
        
        # Create mock database
        client = AsyncMongoMockClient()
        db = client.test_db
        
        comments = [
            CommentCreate(
                comment_text="Test comment",
                source="test"
            )
        ]
        
        custom_upload_id = "custom_upload_456"
        user_id = ObjectId()
        
        # Perform bulk insert with custom ID
        result = await UploadService.bulk_insert_comments(
            comments=comments,
            user_id=user_id,
            db=db,
            upload_id=custom_upload_id
        )
        
        assert result['upload_id'] == custom_upload_id
        
        # Verify progress tracking uses custom ID
        progress = await db.upload_progress.find_one({"upload_id": custom_upload_id})
        assert progress is not None
    
    @pytest.mark.asyncio
    async def test_bulk_insert_progress_percentage(self):
        """Test progress percentage calculation."""
        from app.models.comment import UploadProgress
        
        progress = UploadProgress(
            upload_id="test",
            user_id="user",
            total_comments=100,
            processed_comments=50,
            stored_comments=50,
            failed_comments=0,
            status="processing"
        )
        
        assert progress.progress_percentage == 50.0
        
        progress.processed_comments = 100
        assert progress.progress_percentage == 100.0
        
        # Test zero division
        progress.total_comments = 0
        assert progress.progress_percentage == 0.0
